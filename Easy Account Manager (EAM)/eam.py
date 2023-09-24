import os
import csv
import datetime
import tkinter as tk
from tkinter import ttk
from ttkthemes import ThemedStyle
from tkinter import PhotoImage
from tkinter.ttk import Combobox
from tkinter.ttk import Combobox, Checkbutton, Style
from tkinter import messagebox
from tkinter import filedialog
import json
import pandas as pd
import openpyxl
import getpass  # Módulo para obtener el nombre de usuario de Windows



# Ruta del archivo Excel donde se almacenará el inventario de cuentas
ARCHIVO_INVENTARIO = 'inventario_cuentas.xlsx'  # Cambiado a formato Excel

# Ruta del archivo JSON para los departamentos
ARCHIVO_DEPARTAMENTOS = 'departamentos.json'

# Ruta del archivo JSON para los sistemas
ARCHIVO_SISTEMAS = 'sistemas.json'

# Función para registrar acciones
def registrar_accion(accion, nombre_usuario):
    # Obtiene la hora y la fecha actual
    ahora = datetime.datetime.now()
    
    # Obtiene el nombre de usuario de Windows
    usuario = getpass.getuser()
    
    # Formatea la acción con la hora, fecha y usuario
    registro = f"{ahora:%Y-%m-%d %H:%M:%S} - Usuario: {usuario} - Acción: {accion}"
    
    # Escribe el registro en el archivo de registro
    with open("registro.log", "a") as archivo_log:
        archivo_log.write(registro + "\n")

# Función para cargar los departamentos desde un archivo JSON
def cargar_departamentos():
    if os.path.exists(ARCHIVO_DEPARTAMENTOS):
        with open(ARCHIVO_DEPARTAMENTOS, 'r') as archivo_json:
            departamentos = json.load(archivo_json)
        return departamentos
    else:
        return []

# Función para cargar los sistemas desde un archivo JSON
def cargar_sistemas():
    if os.path.exists(ARCHIVO_SISTEMAS):
        with open(ARCHIVO_SISTEMAS, 'r') as archivo_json:
            sistemas = json.load(archivo_json)
        return sistemas
    else:
        return []

# Función para guardar los departamentos en un archivo JSON
def guardar_departamentos(departamentos):
    with open(ARCHIVO_DEPARTAMENTOS, 'w') as archivo_json:
        json.dump(departamentos, archivo_json)

# Función para guardar los sistemas en un archivo JSON
def guardar_sistemas(sistemas):
    with open(ARCHIVO_SISTEMAS, 'w') as archivo_json:
        json.dump(sistemas, archivo_json)

# Función para crear un nuevo registro de cuenta en el inventario
def crear_registro_cuenta():
    nombre = entry_nombre.get()
    usuario = entry_usuario.get()
    departamento = combo_departamento_var.get()
    sistema = combo_sistema_var.get()  # Nuevo campo para el sistema
    fecha_inicio = entry_fecha_inicio.get()
    fecha_fin = entry_fecha_fin.get()
    privilegio = checkbox_privilegio_var.get()
    servicio = checkbox_servicio_var.get()

    if not os.path.exists(ARCHIVO_INVENTARIO):
        # Si el archivo no existe, se crea y se añade la cabecera
        df = pd.DataFrame(columns=['Nombre', 'Usuario', 'Departamento', 'Sistema', 'Fecha Inicio', 'Fecha Fin', 'Privilegio', 'Servicio'])
    else:
        # Si el archivo existe, cargar los datos actuales
        df = pd.read_excel(ARCHIVO_INVENTARIO)

    # Si el checkbox está marcado, establece la fecha_fin como "Actualmente Trabajando"
    if checkbox_actualmente_trabajando.get():
        fecha_fin = "Actualmente Trabajando"

    # Agregar el nuevo registro al DataFrame
    nuevo_registro = {'Nombre': nombre, 'Usuario': usuario, 'Departamento': departamento, 'Sistema': sistema, 'Fecha Inicio': fecha_inicio, 'Fecha Fin': fecha_fin, 'Privilegio': privilegio, 'Servicio': servicio}
    df = pd.concat([df, pd.DataFrame([nuevo_registro])], ignore_index=True)

    # Guardar el DataFrame en el archivo Excel
    df.to_excel(ARCHIVO_INVENTARIO, index=False)

    # Limpiar los campos después de agregar
    entry_nombre.delete(0, tk.END)
    entry_usuario.delete(0, tk.END)
    combo_departamento_var.set('')
    combo_sistema_var.set('')
    entry_fecha_inicio.delete(0, tk.END)
    entry_fecha_fin.delete(0, tk.END)
    checkbox_actualmente_trabajando.set(0)
    checkbox_privilegio_var.set(0)  # Desmarcar el checkbox de privilegio
    checkbox_servicio_var.set(0)   # Desmarcar el checkbox de servicio
    # Obtener el nombre de usuario actual de Windows
    usuario_actual = getpass.getuser()
    # Registrar la acción en el log, incluyendo el usuario creado
    registrar_accion(f"Creó un usuario ({usuario})", usuario_actual)
    messagebox.showinfo("Registro agregado", "Registro de cuenta agregado con éxito.")

# Función para buscar una cuenta en el inventario por nombre de usuario (username)
def buscar_cuenta_inventario():
    usuario = entry_buscar_usuario.get()
    if os.path.exists(ARCHIVO_INVENTARIO):
        wb = openpyxl.load_workbook(ARCHIVO_INVENTARIO)
        sheet = wb.active
        encontrado = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1] == usuario:
                # Abrir una ventana para modificar el usuario
                modificar_usuario_window(row)
                encontrado = True
        if not encontrado:
            messagebox.showinfo("Cuenta no encontrada", f"No se encontró un registro para el usuario {usuario}.")
    else:
        messagebox.showinfo("Inventario vacío", "El inventario de cuentas está vacío.")

# Función para abrir una ventana para modificar un usuario encontrado
def modificar_usuario_window(usuario_data):
    modificar_usuario_window = tk.Toplevel(root)
    modificar_usuario_window.title("Modificar Usuario")
    modificar_usuario_window.geometry("550x270")

    label_nombre_mod = tk.Label(modificar_usuario_window, text="Nombre Completo:")
    entry_nombre_mod = tk.Entry(modificar_usuario_window)
    entry_nombre_mod.insert(0, usuario_data[0])  # Rellenar con el valor existente

    label_usuario_mod = tk.Label(modificar_usuario_window, text="Nombre de Usuario:")
    entry_usuario_mod = tk.Entry(modificar_usuario_window)
    entry_usuario_mod.insert(0, usuario_data[1])  # Rellenar con el valor existente

    label_departamento_mod = tk.Label(modificar_usuario_window, text="Departamento:")
    combo_departamento_mod_var = tk.StringVar(modificar_usuario_window)  # Variable de control para el departamento
    combo_departamento_mod_var.set(usuario_data[2])  # Establece el valor existente
    combo_departamento_mod_menu = tk.OptionMenu(modificar_usuario_window, combo_departamento_mod_var, *departamentos)

    label_sistema_mod = tk.Label(modificar_usuario_window, text="Sistema:")  # Nuevo campo para el sistema
    combo_sistema_mod_var = tk.StringVar(modificar_usuario_window)  # Variable de control para el sistema
    combo_sistema_mod_var.set(usuario_data[3])  # Establece el valor existente
    combo_sistema_mod_menu = tk.OptionMenu(modificar_usuario_window, combo_sistema_mod_var, *sistemas)
    combo_sistema_mod_menu.config(width=20)  # Ajustar el ancho según sea necesario

    label_fecha_inicio_mod = tk.Label(modificar_usuario_window, text="Fecha de Inicio (YYYY-MM-DD):")
    entry_fecha_inicio_mod = tk.Entry(modificar_usuario_window)
    entry_fecha_inicio_mod.insert(0, usuario_data[4])  # Rellenar con el valor existente

    label_fecha_fin_mod = tk.Label(modificar_usuario_window, text="Fecha de Fin (YYYY-MM-DD o 'Actualmente Trabajando'):")
    entry_fecha_fin_mod = tk.Entry(modificar_usuario_window)
    entry_fecha_fin_mod.insert(0, usuario_data[5])  # Rellenar con el valor existente

    checkbox_actualmente_trabajando_mod_var = tk.BooleanVar()
    checkbox_actualmente_trabajando_mod = tk.Checkbutton(modificar_usuario_window, text="Actualmente Trabajando", variable=checkbox_actualmente_trabajando_mod_var)
    if usuario_data[5] == "Actualmente Trabajando":
        checkbox_actualmente_trabajando_mod.select()
    checkbox_privilegio_mod_var = tk.IntVar()
    checkbox_privilegio_mod = tk.Checkbutton(modificar_usuario_window, text="Cuenta de Privilegio", variable=checkbox_privilegio_mod_var)
    if usuario_data[6]:
        checkbox_privilegio_mod.select()
    checkbox_servicio_mod_var = tk.IntVar()
    checkbox_servicio_mod = tk.Checkbutton(modificar_usuario_window, text="Cuenta de Servicio", variable=checkbox_servicio_mod_var)
    if usuario_data[7]:
        checkbox_servicio_mod.select()

    btn_guardar_modificacion = tk.Button(modificar_usuario_window, text="Guardar Cambios",
                                         command=lambda: guardar_modificacion_usuario(entry_nombre_mod.get(),
                                                                                       entry_usuario_mod.get(),
                                                                                       combo_departamento_mod_var.get(),
                                                                                       combo_sistema_mod_var.get(),
                                                                                       entry_fecha_inicio_mod.get(),
                                                                                       entry_fecha_fin_mod.get(),
                                                                                       checkbox_actualmente_trabajando_mod_var.get(),
                                                                                       checkbox_privilegio_mod_var.get(),
                                                                                       checkbox_servicio_mod_var.get(),
                                                                                       modificar_usuario_window))

    label_nombre_mod.grid(row=0, column=0, sticky="w")
    entry_nombre_mod.grid(row=0, column=1)
    label_usuario_mod.grid(row=1, column=0, sticky="w")
    entry_usuario_mod.grid(row=1, column=1)
    label_departamento_mod.grid(row=2, column=0, sticky="w")
    combo_departamento_mod_menu.grid(row=2, column=1, sticky="w")
    label_sistema_mod.grid(row=3, column=0, sticky="w")  # Posiciona el label del sistema
    combo_sistema_mod_menu.grid(row=3, column=1, sticky="w")  # Posiciona el combo del sistema
    label_fecha_inicio_mod.grid(row=4, column=0, sticky="w")
    entry_fecha_inicio_mod.grid(row=4, column=1)
    label_fecha_fin_mod.grid(row=5, column=0, sticky="w")
    entry_fecha_fin_mod.grid(row=5, column=1)
    checkbox_actualmente_trabajando_mod.grid(row=6, columnspan=2)
    checkbox_privilegio_mod.grid(row=7, columnspan=2)
    checkbox_servicio_mod.grid(row=8, columnspan=2)
    btn_guardar_modificacion.grid(row=9, columnspan=2)

# Función para guardar la modificación de un usuario
def guardar_modificacion_usuario(nombre_mod, usuario_mod, departamento_mod, sistema_mod, fecha_inicio_mod, fecha_fin_mod, actualmente_trabajando_mod, privilegio_mod, servicio_mod, ventana_modificacion):
    if os.path.exists(ARCHIVO_INVENTARIO):
        wb = openpyxl.load_workbook(ARCHIVO_INVENTARIO)
        sheet = wb.active
        encontrado = False
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] == entry_buscar_usuario.get():
                encontrado = True
                # Obtener los valores originales
                nombre_original = row[0]
                usuario_original = row[1]
                departamento_original = row[2]
                sistema_original = row[3]
                fecha_inicio_original = row[4]
                fecha_fin_original = row[5]
                privilegio_original = row[6]
                servicio_original = row[7]
                # Verificar qué campos se modificaron
                cambios = []
                if nombre_original != nombre_mod:
                    cambios.append(f"Nombre: {nombre_original} -> {nombre_mod}")
                if usuario_original != usuario_mod:
                    cambios.append(f"Usuario: {usuario_original} -> {usuario_mod}")
                if departamento_original != departamento_mod:
                    cambios.append(f"Departamento: {departamento_original} -> {departamento_mod}")
                if sistema_original != sistema_mod:
                    cambios.append(f"Sistema: {sistema_original} -> {sistema_mod}")
                if fecha_inicio_original != fecha_inicio_mod:
                    cambios.append(f"Fecha de Inicio: {fecha_inicio_original} -> {fecha_inicio_mod}")
                if fecha_fin_original != fecha_fin_mod:
                    cambios.append(f"Fecha de Fin: {fecha_fin_original} -> {fecha_fin_mod}")
                if privilegio_original != privilegio_mod:
                    cambios.append(f"Privilegio: {privilegio_original} -> {privilegio_mod}")
                if servicio_original != servicio_mod:
                    cambios.append(f"Servicio: {servicio_original} -> {servicio_mod}")
                
                # Verificar el cambio en el estado
                if fecha_fin_original != fecha_fin_mod:
                    cambios.append(f"Fecha de Fin: {fecha_fin_original} -> {fecha_fin_mod}")

                if fecha_fin_original == "Actualmente Trabajando" and not actualmente_trabajando_mod:
                    cambios.append("Estado: Actualmente Trabajando -> No Activo")
                elif fecha_fin_original != "Actualmente Trabajando" and actualmente_trabajando_mod:
                    cambios.append("Estado: No Activo -> Actualmente Trabajando")
                # Actualiza los valores de las celdas
                sheet.cell(row=row_index, column=1, value=nombre_mod)
                sheet.cell(row=row_index, column=2, value=usuario_mod)
                sheet.cell(row=row_index, column=3, value=departamento_mod)
                sheet.cell(row=row_index, column=4, value=sistema_mod)
                sheet.cell(row=row_index, column=5, value=fecha_inicio_mod)
                if actualmente_trabajando_mod:
                    sheet.cell(row=row_index, column=6, value="Actualmente Trabajando")
                    sheet.cell(row=row_index, column=7, value=privilegio_mod)
                    sheet.cell(row=row_index, column=8, value=servicio_mod)
                else:
                    sheet.cell(row=row_index, column=6, value=fecha_fin_mod)
                
                ventana_modificacion.destroy()
                wb.save(ARCHIVO_INVENTARIO)
                # Obtén el nombre de usuario actual
                usuario_actual = getpass.getuser()
                # Registrar la acción en el log, incluyendo los cambios realizados
                registrar_accion(f"Modificó un usuario ({usuario_mod}) - Cambios: {', '.join(cambios)}", usuario_actual)
                messagebox.showinfo("Usuario modificado", "Los cambios han sido guardados.")
                break
        
        if not encontrado:
            messagebox.showinfo("Usuario no encontrado", "No se encontró un registro para el usuario buscado.")
    else:
        messagebox.showinfo("Inventario vacío", "El inventario de cuentas está vacío.")

# Función para eliminar una cuenta del inventario por nombre de usuario (username)
def eliminar_cuenta_inventario():
    usuario = entry_eliminar_usuario.get()
    if os.path.exists(ARCHIVO_INVENTARIO):
        wb = openpyxl.load_workbook(ARCHIVO_INVENTARIO)
        sheet = wb.active
        eliminado = False
        for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row[1] == usuario:
                # Elimina la fila que contiene el usuario
                sheet.delete_rows(row_index)
                eliminado = True
                break
        if eliminado:
            wb.save(ARCHIVO_INVENTARIO)
            messagebox.showinfo("Cuenta eliminada", f"El registro del usuario {usuario} ha sido eliminado del inventario.")
            # Obtener el nombre de usuario actual de Windows
            usuario_actual = getpass.getuser()
            # Registrar la acción en el log, incluyendo el usuario creado
            registrar_accion(f"Eliminó un usuario ({usuario})", usuario_actual)
        else:
            messagebox.showinfo("Cuenta no encontrada", f"No se encontró un registro para el usuario {usuario}.")
    else:
        messagebox.showinfo("Inventario vacío", "El inventario de cuentas está vacío.")


# Función para abrir la ventana de ajustes de departamentos
def abrir_ventana_ajustes():
    ajustes_window = tk.Toplevel(root)
    ajustes_window.title("Ajustes de Departamentos")
    ajustes_window.geometry("400x100")
    
    # Crear y cargar la lista de departamentos desde el archivo JSON
    departamentos = cargar_departamentos()
    
    def guardar_departamentos_ajustes():
        nuevos_departamentos = entry_nuevos_departamentos.get().split(',')
        nuevos_departamentos = [dep.strip() for dep in nuevos_departamentos]
        
        # Agregar los nuevos departamentos a la lista existente
        departamentos.extend(nuevos_departamentos)
        
        # Guardar la lista actualizada en el archivo JSON
        guardar_departamentos(departamentos)
        
        # Actualizar la lista desplegable en la ventana principal
        # Actualizar la lista desplegable en la ventana principal
        combo_departamento_var.set('')  # Limpiar la selección actual
        combo_departamento_menu['values'] = departamentos  # Actualizar la lista de valores
        
        messagebox.showinfo("Departamentos actualizados", "La lista de departamentos ha sido actualizada con éxito.")
        ajustes_window.destroy()
    
    label_ajustes = tk.Label(ajustes_window, text="Agregar Nuevos Departamentos (separados por coma):")
    entry_nuevos_departamentos = tk.Entry(ajustes_window)
    btn_guardar_departamentos = tk.Button(ajustes_window, text="Guardar Departamentos", command=guardar_departamentos_ajustes)
    
    label_ajustes.pack()
    entry_nuevos_departamentos.pack()
    btn_guardar_departamentos.pack()

# Función para abrir la ventana de ajustes de sistemas
def abrir_ventana_ajustes_sistemas():
    ajustes_sistemas_window = tk.Toplevel(root)
    ajustes_sistemas_window.title("Ajustes de Sistemas")
    ajustes_sistemas_window.geometry("400x100")
    
    # Crear y cargar la lista de sistemas desde el archivo JSON
    sistemas = cargar_sistemas()
    
    def guardar_sistemas_ajustes():
        nuevos_sistemas = entry_nuevos_sistemas.get().split(',')
        nuevos_sistemas = [sistema.strip() for sistema in nuevos_sistemas]
        
        # Agregar los nuevos sistemas a la lista existente
        sistemas.extend(nuevos_sistemas)
        
        # Guardar la lista actualizada en el archivo JSON
        guardar_sistemas(sistemas)
        
        # Actualizar la lista desplegable en la ventana principal
        combo_sistema_var.set('')
        combo_sistema_var.set('')  # Limpiar la selección actual
        combo_sistema_menu['values'] = sistemas  # Actualizar la lista de valores
        
        messagebox.showinfo("Sistemas actualizados", "La lista de sistemas ha sido actualizada con éxito.")
        ajustes_sistemas_window.destroy()
    
    label_ajustes_sistemas = tk.Label(ajustes_sistemas_window, text="Agregar Nuevos Sistemas (separados por coma):")
    entry_nuevos_sistemas = tk.Entry(ajustes_sistemas_window)
    btn_guardar_sistemas = tk.Button(ajustes_sistemas_window, text="Guardar Sistemas", command=guardar_sistemas_ajustes)
    
    label_ajustes_sistemas.pack()
    entry_nuevos_sistemas.pack()
    btn_guardar_sistemas.pack()

# Crear la ventana principal de la aplicación
root = tk.Tk()
root.title("Easy Account Manager - by Damian de Arce")
root.geometry("600x810")

# Estilo de ttkthemes
style = ThemedStyle(root)
style.set_theme("breeze")  # Cambia el tema a uno más moderno (puedes elegir otro si prefieres)

# Agregar íconos a los botones
add_icon = PhotoImage(file='add.png')
search_icon = PhotoImage(file='search.png')
delete_icon = PhotoImage(file='delete.png')
setting_icon = PhotoImage(file='setting.png')

# Crear y cargar la lista de departamentos desde el archivo JSON
departamentos = cargar_departamentos()

# Crear y cargar la lista de sistemas desde el archivo JSON
sistemas = cargar_sistemas()

# Caja "Ingreso de cuenta"
frame_ingreso_cuenta = ttk.LabelFrame(root, text="Ingreso de cuenta", padding=(10, 10))
frame_ingreso_cuenta.pack(fill="both", padx=20, pady=10)

# Widgets dentro de la caja "Ingreso de cuenta"
label_nombre = ttk.Label(frame_ingreso_cuenta, text="Nombre Completo:")
entry_nombre = ttk.Entry(frame_ingreso_cuenta)
label_usuario = ttk.Label(frame_ingreso_cuenta, text="Nombre de Usuario:")
entry_usuario = ttk.Entry(frame_ingreso_cuenta)
label_departamento = ttk.Label(frame_ingreso_cuenta, text="Departamento:")
combo_departamento_var = tk.StringVar(root)
combo_departamento_var.set('')
combo_departamento_menu = Combobox(frame_ingreso_cuenta, textvariable=combo_departamento_var, values=departamentos)
label_sistema = ttk.Label(frame_ingreso_cuenta, text="Sistema:")
combo_sistema_var = tk.StringVar(root)
combo_sistema_var.set('')
combo_sistema_menu = Combobox(frame_ingreso_cuenta, textvariable=combo_sistema_var, values=sistemas)
label_fecha_inicio = ttk.Label(frame_ingreso_cuenta, text="Fecha de Inicio (DD-MM-YYYY):")
entry_fecha_inicio = ttk.Entry(frame_ingreso_cuenta)
label_fecha_fin = ttk.Label(frame_ingreso_cuenta, text="Fecha de Fin (DD-MM-YYYY o 'Actualmente Trabajando'):")
entry_fecha_fin = ttk.Entry(frame_ingreso_cuenta)
checkbox_actualmente_trabajando = tk.IntVar()
checkbox_actualmente_trabajando.set(0)
checkbox_actualmente_trabajando_checkbutton = Checkbutton(frame_ingreso_cuenta, text="Actualmente Trabajando", variable=checkbox_actualmente_trabajando)
checkbox_privilegio_var = tk.IntVar()
checkbox_privilegio = tk.Checkbutton(frame_ingreso_cuenta, text="Cuenta de Privilegio", variable=checkbox_privilegio_var)
checkbox_servicio_var = tk.IntVar()
checkbox_servicio = tk.Checkbutton(frame_ingreso_cuenta, text="Cuenta de Servicio", variable=checkbox_servicio_var)
btn_agregar = ttk.Button(frame_ingreso_cuenta, text="Agregar Usuario al Inventario", command=crear_registro_cuenta, image=add_icon, compound="left")

# Posicionamiento de widgets en la caja "Ingreso de cuenta"
label_nombre.grid(row=0, column=0, sticky="w")
entry_nombre.grid(row=0, column=1)
label_usuario.grid(row=1, column=0, sticky="w")
entry_usuario.grid(row=1, column=1)
label_departamento.grid(row=2, column=0, sticky="w")
combo_departamento_menu.grid(row=2, column=1, sticky="w")
label_sistema.grid(row=3, column=0, sticky="w")
combo_sistema_menu.grid(row=3, column=1, sticky="w")
label_fecha_inicio.grid(row=4, column=0, sticky="w")
entry_fecha_inicio.grid(row=4, column=1)
label_fecha_fin.grid(row=5, column=0, sticky="w")
entry_fecha_fin.grid(row=5, column=1)
checkbox_actualmente_trabajando_checkbutton.grid(row=6, columnspan=2)
checkbox_privilegio.grid(row=7, columnspan=2)
checkbox_servicio.grid(row=8, columnspan=2)
btn_agregar.grid(row=9, columnspan=2)

# Caja "Modificar cuenta"
frame_modificar_cuenta = ttk.LabelFrame(root, text="Modificar cuenta", padding=(10, 10))
frame_modificar_cuenta.pack(fill="both", padx=20, pady=10)

# Widgets dentro de la caja "Modificar cuenta"
label_buscar_usuario = ttk.Label(frame_modificar_cuenta, text="Buscar Usuario por Nombre:")
entry_buscar_usuario = ttk.Entry(frame_modificar_cuenta)
btn_buscar_usuario = ttk.Button(frame_modificar_cuenta, text="Buscar Usuario", command=buscar_cuenta_inventario, image=search_icon, compound="left")

label_eliminar_usuario = ttk.Label(frame_modificar_cuenta, text="Eliminar Usuario por Nombre:")
entry_eliminar_usuario = ttk.Entry(frame_modificar_cuenta)
btn_eliminar_usuario = ttk.Button(frame_modificar_cuenta, text="Eliminar Usuario", command=eliminar_cuenta_inventario, image=delete_icon, compound="left")

# Posicionamiento de widgets en la caja "Modificar cuenta"
label_buscar_usuario.pack(anchor="w", padx=(10, 25), pady=(10, 0))
entry_buscar_usuario.pack(padx=10, pady=(0, 10))
btn_buscar_usuario.pack(fill="none", padx=10, pady=(0, 10))
label_eliminar_usuario.pack(anchor="w", padx=(10, 0), pady=(10, 0))
entry_eliminar_usuario.pack(padx=10, pady=(0, 10))
btn_eliminar_usuario.pack(fill="none", padx=10, pady=(0, 10))

# Caja "Ajustes"
frame_ajustes = ttk.LabelFrame(root, text="Ajustes", padding=(10, 10))
frame_ajustes.pack(fill="both", padx=20, pady=10)

# Widgets dentro de la caja "Ajustes"
btn_ajustes_departamentos = ttk.Button(frame_ajustes, text="Ajustes de Departamentos", command=abrir_ventana_ajustes, image=setting_icon, compound="left")
btn_ajustes_sistemas = ttk.Button(frame_ajustes, text="Ajustes de Sistemas", command=abrir_ventana_ajustes_sistemas, image=setting_icon, compound="left")

# Posicionamiento de widgets en la caja "Ajustes"
btn_ajustes_departamentos.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")
btn_ajustes_sistemas.grid(row=0, column=1, padx=10, pady=10, sticky="nsew")
frame_ajustes.columnconfigure(0, weight=1)
frame_ajustes.columnconfigure(1, weight=1)
frame_ajustes.rowconfigure(0, weight=1)

root.mainloop()
